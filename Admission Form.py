from tkinter import *
from tkcalendar import *
import tkinter as tk
from tkinter import ttk
import openpyxl
from openpyxl import Workbook
import pathlib

main = Tk()
main.title("ADMISSION FORM")
main.geometry("1000x700")

file = pathlib.Path("Backends.xlsx") #Generating File
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet["A1"] = "Full Name"
    sheet["B1"] = "Email"
    sheet["C1"] = "Gender"
    sheet["D1"] = "D.O.B"
    sheet["E1"] = "Class 10"
    sheet["F1"] = "Class 12"
    sheet["G1"] = "UNIVERSITY"
    sheet["H1"] = "COLLEGE"
    sheet["I1"] = "DEPT"
    sheet["J1"] = "Mobile No"

    file.save("Backends.xlsx")

# frame 1 structure
frame1 =LabelFrame(main,text="Personal Details").pack(expand='yes',fill='both')

Label(frame1,text='Name').place(x=50,y=30)
Label(frame1,text='Email id').place(x=50,y=60)
Label(frame1,text='Gender').place(x=50,y=90)
Label(frame1,text='Mobile No').place(x=50,y=150)

#Function to generate the Calendar

def show():
    top=Toplevel(main)
    def grabdate():
        global dob_s
        global dob
        dob_s = StringVar()
        dob = tk.Label(frame1, text=cal.get_date())
        dob.pack()
        dob.place(x=150, y=120)
        dob_s=cal.get_date()


    cal=Calendar(top,setmode='day',date_pattern='dd/mm/yyyy')
    cal.pack()

    sel = Button(top,text='select date',command=grabdate).pack()

d_B=Button(frame1,text='Date of Birth',command=show).place(x=50,y=120)
dob_s=tk.StringVar()

#Name Variable
name_s=tk.StringVar()
name_enter=ttk.Entry(frame1,width=25,textvariable=name_s)
name_enter.place(x=150,y=30)
name_enter.focus()

#Email Variable
email_s=tk.StringVar()
email_enter=ttk.Entry(frame1,width=30,textvariable=email_s)
email_enter.place(x=150,y=60)
email_enter.focus()

#Gender Variable
gender_s=tk.StringVar()
gender_combobox=ttk.Combobox(frame1,width=10,textvariable=gender_s,state="readonly")#state make user to only select from given option
gender_combobox["values"]=("MALE","FEMALE")
gender_combobox.current(0)#for showing a default value
gender_combobox.place(x=150,y=90)

#Mob No var
MobNo_s=StringVar()
MobNo_enter=ttk.Entry(frame1,width=10,textvariable=MobNo_s)
MobNo_enter.place(x=150,y=150)

#Frame2 Structure
frame2 =LabelFrame(main,text="Academic Details").pack(expand='yes',fill='both')
Label(frame2,text='Class 10 board').place(x=50,y=250)
Label(frame2,text='Class 12 board').place(x=50,y=280)
Label(frame2,text='Passing Year').place(x=350,y=250)
Label(frame2,text='Passing Year').place(x=350,y=280)
Label(frame2,text='Marks Obtain').place(x=550,y=250)
Label(frame2,text='Marks Obtain').place(x=550,y=280)

board_s=tk.StringVar()
board_combobox=ttk.Combobox(frame2,width=15,textvariable=board_s,state="readonly")#state make user to only select from given option
board_combobox["values"]=("WBSE","CBSE","ICSE","OTHERS")
board_combobox.current(0)#for showing a default value
board_combobox.place(x=180,y=250)

board2_s=tk.StringVar()
board2_combobox=ttk.Combobox(frame2,width=15,textvariable=board2_s,state="readonly")#state make user to only select from given option
board2_combobox["values"]=("WBCHSE","CBSE","ICSE","OTHERS")
board2_combobox.current(0)  #for showing a default value
board2_combobox.place(x=180,y=280)

pyear_s=tk.StringVar()
pyear_spinbox=ttk.Spinbox(frame2,width=5,textvariable=pyear_s,from_=2010,to_=2019)#state make user to only select from given option
pyear_spinbox.place(x=450,y=250)

pyear2_s=tk.StringVar()
pyear2_spinbox=ttk.Spinbox(frame2,width=5,textvariable=pyear2_s,from_=2012,to_=2021)#state make user to only select from given option
pyear2_spinbox.place(x=450,y=280)

m1_s=tk.StringVar()
m1_enter=ttk.Entry(frame2,width=5,textvariable=m1_s)
m1_enter.place(x=650,y=250)


m2_s=tk.StringVar()
m2_enter=ttk.Entry(frame2,width=5,textvariable=m2_s)
m2_enter.place(x=650,y=280)

#Frame 3 Structure
frame3 =LabelFrame(main,text="Institution Details").pack(expand='yes',fill='both')
Label(frame3,text="University Name").place(x=50,y=450)
Label(frame3,text="College Name").place(x=50,y=480)
Label(frame3,text="Department").place(x=50,y=510)

u_s=tk.StringVar()
u_enter=ttk.Entry(frame2,width=20,textvariable=u_s)
u_enter.place(x=150,y=450)
u_enter.focus()

col_s=tk.StringVar()
col_enter=ttk.Entry(frame2,width=20,textvariable=col_s)
col_enter.place(x=150,y=480)
col_enter.focus()

dept_s=tk.StringVar()
dept_combobox=ttk.Combobox(frame2,width=15,textvariable=dept_s,state="readonly")#state make user to only select from given option
dept_combobox["values"]=("CSE","ME","EE","ECE","CIVIL","OTHERS")
dept_combobox.current(0)#for showing a default value
dept_combobox.place(x=150,y=510)

# Submit & Reset Button
def submit():
    a=name_s.get()
    b=email_s.get()
    c=gender_s.get()
    d=dob_s
    e="Board: "+board_s.get()
    f="Passed In "+pyear_s.get()
    g="Marks: "+m1_s.get()
    h="Board: "+board2_s.get()
    i="Passed In "+pyear2_s.get()
    j="Marks: "+m2_s.get()
    k=u_s.get()
    l=col_s.get()
    m=dept_s.get()
    n=MobNo_s.get()
    print(a,b,c,d,e,f,g,h,i,j,k,l,m,n,sep='\n')

    file = openpyxl.load_workbook("Backends.xlsx")
    sheet=file.active
    sheet.cell(column=1, row=sheet.max_row+2, value=a)
    sheet.cell(column=2, row=sheet.max_row, value=b)
    sheet.cell(column=3, row=sheet.max_row, value=c)
    sheet.cell(column=4, row=sheet.max_row, value=d)
    sheet.cell(column=5, row=sheet.max_row, value=e)
    sheet.cell(column=5, row=sheet.max_row+1, value=f)
    sheet.cell(column=5, row=sheet.max_row+1, value=g)
    sheet.cell(column=6, row=sheet.max_row - 2, value=h)
    sheet.cell(column=6, row=sheet.max_row - 1, value=i)
    sheet.cell(column=6, row=sheet.max_row, value=j)
    sheet.cell(column=7, row=sheet.max_row - 2, value=k)
    sheet.cell(column=8, row=sheet.max_row - 2, value=l)
    sheet.cell(column=9, row=sheet.max_row - 2, value=m)
    sheet.cell(column=10, row=sheet.max_row - 2, value=n)


    file.save("Backends.xlsx")

def reset():
    name_enter.delete(0,tk.END)
    email_enter.delete(0, tk.END)
    u_enter.delete(0, tk.END)
    col_enter.delete(0,tk.END)
    m1_enter.delete(0,tk.END)
    m2_enter.delete(0,tk.END)
    MobNo_enter.delete(0, tk.END)
    gender_combobox.current(0)
    board_combobox.current(0)
    board2_combobox.current(0)
    dept_combobox.current(0)
    pyear_s=StringVar()
    pyear2_s=StringVar()
    pyear_spinbox.config(textvariable=pyear_s)
    pyear2_spinbox.config(textvariable=pyear2_s)
    dob.destroy()


btn1=Button(main,text="SUBMIT",width=6,bg='green',command=submit).pack()
btn2=Button(main,text="RESET",width=6,bg='yellow',command=reset).pack()

main.mainloop()

